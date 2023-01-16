def extract_diameters(sh) -> list:
    """Takes in Curve PSD worksheet (openpyxl), returns list of all the trim diameters from cells A10-A21"""    
    list_of_diams = []    
    for row in sh.iter_rows(min_row=10, max_col=1, max_row=21):
        for cell in row:
            if cell.value != None:
                list_of_diams.append(cell.value)
    return(list_of_diams)


def reset_column_A(diameter, sh, diameters_list):
    """Sets only 1 diameter in diameter column (Col A)"""
    for index, value in enumerate(diameters_list):
        cell_to_change = "{}{}".format('A', 10+index)
        cell_to_change2 = "{}{}".format('B', 10+index)
        
        if index == 0:
            sh[cell_to_change].value = diameter
            sh[cell_to_change2].value = str(diameter) + '-' + str(diameter)
        else:
            sh[cell_to_change].value = ''
            sh[cell_to_change2].value = ''   
    return


def removeCurveData(diameter, sh, list_of_trims):
    """ Deletes curve data not pertaining to current diameter for the tab"""    
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter

    first_col = 4
    max_cols = 21 * len(list_of_trims)
    row = 7    

    diameter_cols_idx = sorted(list(range(first_col,max_cols,21)), reverse=True)           # Assuming diameters are listed from largest to smallest, here we locate where curve data is for other trims.
   
    for col_index in diameter_cols_idx:
        cell = get_column_letter(col_index) + str(row)
        
        if sh[cell].value != diameter:                                                     # Check diameter values in row 7. IF diameter doesn't match, delete the following 21 columns
            sh.delete_cols(col_index, 21)
    return

if __name__ == "__main__":
    ''' 1. Load excel
        2. Find only tabs with "P"
        3. Count how many diameters (curve tables) there are in Column A
            a. Make that many duplicate tabs
            b. Rename tabs to include Diameter
            c. Delete the unnecessary tables that don't match the diameter for the tab'  
        4. save to excel file '''

    import os
    from openpyxl import load_workbook
    import shutil

    ########################################################################
    # NEED TO MODIFY FOLDERS, FILENAMES BELOW TO REFLECT FOLDER SETUP
    ########################################################################
    myDir = r"C:\Users\104092\OneDrive - Grundfos\Documents\10-19 Projects\12 NBS Curve PSD Separation\12.01 Original Files"
    myFile = "GXS Curve_Conexus_V2.xlsx"
    original = os.path.join(myDir, myFile)

    filename, file_ext = myFile.split(os.extsep)

    targetDir = r"C:\Users\104092\OneDrive - Grundfos\Documents\10-19 Projects\12 NBS Curve PSD Separation\12.02 Output Files"
    std_target = os.path.join(targetDir, filename + " - std models." + file_ext)
    ########################################################################
    shutil.copyfile(original, std_target)                                               # Create copy to leave original untouched

    wb_std = load_workbook(std_target, data_only=True)

    trim_count = 0

    for sheet in wb_std.worksheets:
        if '-E' in sheet.title:                                                         # Remove tabs ending in -E. Do not want E models
            wb_std.remove(sheet)    

        elif sheet.title[-1] == 'P':                                                    # If tab with "P" is found at end of title, expand the trims found there.
            diameters_list = extract_diameters(sheet)                                   # Extract all trims from the current sheet
        
            for eachDiameter in diameters_list:                                         # Iterate through each trim found in sheet
                trim_count += 1                                                         # Keep track of number of tabs created
                
                newTabName = sheet.title + '-' + str(eachDiameter) + '_Std'             # Add as many tabs as needed per diameter and rename to reflect diameter
                tab_to_copy = wb_std[sheet.title]  
                wb_std.copy_worksheet(tab_to_copy).title = newTabName
                
                reset_column_A(eachDiameter, wb_std[newTabName], diameters_list)        # Remove all additional diameters in column A
                removeCurveData(eachDiameter, wb_std[newTabName], diameters_list)       # Delete data that doesn't belong in the trim diameter tab
            
            wb_std.remove(sheet)                                                        # This removes the original tab 

    print(f"Total Amount of Trims Found: {trim_count}")    
        
    wb_std._sheets.sort(key=lambda ws: ws.title)                                        # This sorts the new excel tabs by name.
    wb_std.save(std_target)