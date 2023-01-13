# -*- coding: utf-8 -*-
"""
Created on Mon Mar 21 09:33:46 2022
Finished: 5/19/22

HOW TO USE
1. Update directory and filepath to pump curve export from GPC (csv)
2. Set output path for PSD

"""
##############################################################################
"""Creates dictionary of dataframes for each PN """
def curveDataByPartNumber(data):
    import math
    
    # This will become a dictionary of dataframes of pns and curve data
    dict_curveDataByPn = {}
    
    # Iterate through each row in pump curve data export 
    for index, value in data.iterrows():    
        # When a PN is encountered:
        if not math.isnan(value['ProductNumber']):
            currentProductNumber = int(value['ProductNumber'])
            # Resets list for each PN
            listOfFlows = []
            listOfHeads = []
            listOfEff = []
            listOfSpeeds = []
            listOfNPSH = []
            
        # Recording Q/H values to lists
        if value['RPM(Curve nominal)'] > 0:
            listOfFlows.append(value['Q'])
            listOfHeads.append(value['H'])
            listOfEff.append(value['Eta1'])
            listOfSpeeds.append(value['RPM(Curve nominal)'])
            listOfNPSH.append(value['NPSH'])
            
        # At end of Q/H values, store in dataframe before moving to next PN
        if pd.isna(value['Q']) and pd.isna(value['H']):    
            zipped = list(zip(listOfFlows, listOfHeads, listOfEff, listOfSpeeds, listOfNPSH))
            df = pd.DataFrame(zipped, columns=['Q','H','Eta1','RPM','NPSH'])
            # Drop rows that have NaNs, then add df to dictionary
            df = df.dropna()
            dict_curveDataByPn.update({currentProductNumber:df})
            
            continue
          
    return dict_curveDataByPn
###############################################################################      
""" Create list of PNs that share curve data """
def curveDataMatches(curveDataDict):  
    
    curveDataCopy = curveDataDict.copy()
    listOfLists = []
    
    for eachPN, eachDF in curveDataDict.items():  
        listOfPNs = []
        for key, value in curveDataCopy.items():
            if value.equals(eachDF):
                listOfPNs.append(key)
        
        if listOfPNs not in listOfLists:
            listOfLists.append(listOfPNs)
            
    return listOfLists
#################################################################################
""" Create list of PNs that share curve data, and inserts tab to illustrate """
def addCurveFamiliesTab():
    
    global wb    
    curveFamilySheet = wb.create_sheet("Shared Curves", 2)
    
    for index, family in enumerate(uniqueCurvePartnumbers):
        column = index + 1
        curveFamilySheet.cell(1, column).value = "Curve " + str(column)
        for ind, partnumber in enumerate(family):
            row = ind + 2
            curveFamilySheet.cell(row, column).value = partnumber

    return
##############################################################################
""" CREATE new tabs for each unique curve family """
def add_a_curve(list_unique_curves):  
    
    global wb
    
    # Add a tab for each unique curve
    for item in list_unique_curves: 
        
        list_of_speeds = []
        
        # Add 1 tab
        tabName = str(item[0])
        wb.copy_worksheet(tab_to_copy).title = tabName
        
        # Fill in all speeds in column A
        for key, value in curveDataDict[int(tabName)].iterrows():
            list_of_speeds.append(value['RPM'])
        
        speed_set = sorted(set(list_of_speeds), reverse=True)
        print(tabName, speed_set)  
        
        for index, eachSpeed in enumerate(speed_set):
            cell_name = "{}{}".format('A', 10+index)
            curveSheet = wb[tabName]
            curveSheet[cell_name].value = int(eachSpeed)
       
    return()
##############################################################################
""" function to find which cell to start populating curve data based on RPM"""
def findSpeedCells(my_df):
    
    # from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
    from openpyxl.utils.cell import get_column_letter
    speedCells = []
    
    max_cols = 21 * len(my_df.RPM.unique())
    row = 7    
    first_col = 4
    diameter_cols = list(range(first_col,max_cols,21))
    
    for item in diameter_cols:
        col = get_column_letter(item)
        cell_coordinate = "{}{}".format(col,row)
        speedCells.append(cell_coordinate)      

    return(speedCells)
#############################################################################
""" fills curve tables in each tab """
def curveFiller():

    global wb
     
    # Iterate through each tab in workbook
    for sheet in wb.worksheets:
        try:
            sheetname = int(sheet.title)
        except ValueError: # This skips non curve data tabs
            continue    
        
        # If tab name is a match in curvedata dictionary, use that dictionary to fill cells with data
        if sheetname in curveDataDict.keys():
            df = curveDataDict[sheetname]
            RPM_cells = findSpeedCells(df)
            print("Curve: {}, RPM cells at: {}".format(sheet.title, RPM_cells))
    
            # Extracts RPMs from D7, Y7, AT7, etc...
            for cellName in RPM_cells:
                # print("current RPM cell: {}".format(cellName))
                cell_reference = sheet[cellName].value[1:]
                rpm = sheet[cell_reference].value
                # print("searching for dataframe rows w/ RPM = {}".format(rpm))
                                
                # This creates sub-dataframes, each with one RPM
                speed_specific_df = df[df["RPM"] == rpm].reset_index()
                
                # Iterate through dataframe row by row and fill out each row in PSD    
                for key, value in speed_specific_df.iterrows():
                    first_row_offset = 3
                    sheet[cellName].offset(first_row_offset + key, 0).value = value['Q']
                    sheet[cellName].offset(first_row_offset + key, 1).value = value['H']
                    sheet[cellName].offset(first_row_offset + key, 7).value = value['Q']
                    sheet[cellName].offset(first_row_offset + key, 8).value = value['Eta1']
                    sheet[cellName].offset(first_row_offset + key,14).value = value['Q']
                    sheet[cellName].offset(first_row_offset + key,15).value = value['NPSH']      
    return   
###############################################################################
"""
        SETUP 
"""
###############################################################################
import os
import pandas as pd
import shutil
from openpyxl import load_workbook

# This is the folder/file with the curve export csv
myDir = r"C:\Users\104092\OneDrive - Grundfos\Documents\10-19 Projects\12 NBS Curve PSD Separation\12.01 Original Files"
myFile = "GPC NBS Curves.xlsx"
filePath = os.path.join(myDir, myFile)

# This creates a dataframe of the curve export csv, and fills in the RPM(curve nominal) column
data = pd.read_csv(filePath, sep=";", index_col=False, skip_blank_lines=False)
data['RPM(Curve nominal)'] = data['RPM(Curve nominal)'].ffill()

# This points to the curve PSD template to be used
templateDir = r"C:\Users\104092\OneDrive - Grundfos\Documents\30-39 Resources\32 GXS"
template = "SKB Blank Curve PSD - Efficiency_Metric.xlsx"
template = os.path.join(templateDir, template)

# Create a local working copy to leave template unmodified
workingCopy = os.path.join(myDir, "Test Curve Template.xlsx")
shutil.copyfile(template, workingCopy)
wb = load_workbook(workingCopy)      
tab_to_copy = wb['NEW']    

# Creates dictionary with part numbers as keys, curves as dataframes for each key
curveDataDict = curveDataByPartNumber(data)    

# Creates list of which part numbers share curve data
uniqueCurvePartnumbers = curveDataMatches(curveDataDict)    

# Adds curve tabs for each unique curve
add_a_curve(uniqueCurvePartnumbers)

# Add tab to show which partnumbers share curves
addCurveFamiliesTab()

# Fills each new curve tab
curveFiller()

# Save changes to excel sheet
wb.save(workingCopy)
###############################################################################
