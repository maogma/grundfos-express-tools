import os
from utils.file_ops import add_filename_timestamp
import shutil
import datetime
import pandas as pd
import numpy as np
from openpyxl import load_workbook

myDir = r"C:\Users\104092\OneDrive - Grundfos\Documents\git\grundfos-express-tools\bronze impeller removal\input files"

# Create working copy
completed_dir = r"C:\Users\104092\OneDrive - Grundfos\Documents\git\grundfos-express-tools\bronze impeller removal\output files"
# working_copy = os.path.join(completed_dir, add_filename_timestamp(file))
# shutil.copyfile(filePath, working_copy) # creates working copy to leave original file untouched

### List of partnumbers to be removed

list_of_pns = [
    96699290, 97775274, 96699299, 97775277, 96778078,
    97780992, 96699305, 96769184, 97778033, 96769190,
    96769205, 97778039, 96769256, 96896891, 96769259,
    96769271, 97780970, 96769280, 97780973
]

str_list = [str(numstr) for numstr in list_of_pns]

### Creating justification note to add to data 
# Reason/Justification for removal
timeStamp = datetime.datetime.now().strftime("%m-%d-%Y")
reason = input("Reason for removal: ")
authority = input("Who authorized/requested this change? ")

removal_note = timeStamp + " " + reason + " per " + authority

### Reading in excel to dataframe

# Customize based on PST template used. 
psd_startrow = 6 # Corresponds to row number containing "second" column headers
psd_startcol = 0

def group_then_separate_by(list_of_cols: list, pn_col: str):
    """list of cols are grouping categories. pn_col is the column that contains pns to find matches on."""
    groups = psd_data.groupby(list_of_cols) # Had to play around with this to get the right groupings

    df_list_to_remove = [] # will hold list of dataframes to be removed. Will concatenate to 1 dataframe later
    df_list_to_keep   = [] # will hold list of dataframes to remain in PSD. Will concatenate to 1 dataframe later

    # Iterates through each sub-group, checking if there is a pn that meets criteria for removal
    for _, frame in groups:
        if any(frame[pn_col].isin(str_list)):   # Finding matching partnumbers to remove
            df_list_to_remove.append(frame)     # Need to add this sub-group to a "removed dataframe"
        else:
            df_list_to_keep.append(frame)       # Need to retain this sub-group, add to a "keep dataframe"

    # Concatenating list of dfs to single dfs.
    removals = pd.concat(df_list_to_remove)
    keep = pd.concat(df_list_to_keep)
        
    return removals, keep

""" Iterates through all files within a directory and performs function on each file"""
for root, subdirectories, files in os.walk(myDir):
    for file in files:
        # fileName, fileExtension = os.path.splitext(file)
        inputFile = os.path.join(myDir, file)
        outputPath = os.path.join(completed_dir, file)
        shutil.copyfile(inputFile, outputPath) # creates working copy to leave original file untouched

        # Read in Excel sheet to a dataframe
        sheetname = "Impeller"
        psd_data = pd.read_excel(outputPath, sheet_name=sheetname, header=psd_startrow-1, dtype={'BOM': str})
        original_psd_bottom_row = len(psd_data) + psd_startrow  

        print("Opening file for updates: {}".format(file))

        ### Grouping and separating data

        # Need to find row number of [END] in PSD
        end_row = psd_data[psd_data['Full Data'] == '[END]'].index.to_list()[0]

        # This removes all the trash removals from the original PSD
        psd_data = psd_data.iloc[:end_row]

        # Find matching partnumber's from list above, and then capture "process variants"
        group_by_columns = ["Model", "Price ID"]  # For just one column, use ["Col 1", ]
        removals, keep = group_then_separate_by(group_by_columns, "BOM")

        # Dataframe formatting.
        removals.loc[removals["Full Data"] == "[START]", "Full Data"] = "" # This removes the [START] if the first PSD row is flagged to be removed

        # Insert removal justification note to top of removal dataframe
        new_row = pd.DataFrame({'ID': removal_note}, index =[0])
        removals = pd.concat([new_row, removals[:]]).reset_index() # Concatenate new_row with df

        column_list = keep.columns
        removals = removals[column_list] # Reorder columns to match the 2 dataframe columns prior to concatenation

        removals.sort_values(by=['ID'], inplace=True)
        # removals.head(10)

        keep.reset_index(drop=True, inplace=True)
        keep.loc[0,'Full Data'] = "[START]"

        ### Checks for any misses
        # Check for any misses before writing. Should return empty dataframe
        if (len(keep[keep['BOM'].isin(str_list)])) == 0:
            print("Didn't miss any partnumbers. Good to go")
        else:
            print("For some reason, the following entries were missed")
            print(keep[keep['BOM'].isin(str_list)])

        # Create copy of PSD tab with formatting
        wb = load_workbook(outputPath)
        ws = wb['Impeller']
        tabName = sheetname + " No Bronze"
        wb.copy_worksheet(ws).title = tabName

        # Find row number to append removals and insert reason why they were removed 
        append_location = original_psd_bottom_row - len(removals) + 1     # This is the new location/row where we will append the removals to bottom of PSD

        # Edit New Tab to clear cells with old data after first [END] value is reached
        num_rows = len(keep)                          # Finds where new [END] row will be on PSD
        after_end_row = num_rows + psd_startrow + 1   # Calculates where [END] row + 1 will be in the PSD

        ws_modified = wb[tabName]
        ws_modified.delete_rows(after_end_row, len(removals)-1)   # clears the rows containing old data

        # Fix formatting
        from openpyxl.styles import PatternFill
        for rows in ws_modified.iter_rows(min_row=append_location+2, max_row=append_location+2, min_col=1, max_col=40):
            for cell in rows:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        wb.save(outputPath)

        # Write resulting dataframes to new sheet in PSD with changes.
        with pd.ExcelWriter(outputPath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:  
            keep.to_excel(writer, sheet_name=tabName, index=False, startrow=psd_startrow-1, startcol=psd_startcol)
            removals.to_excel(writer, sheet_name=tabName, index=False, header=False, startrow=append_location+1, startcol=psd_startcol)

        # Remove original tab, replace with new one?



