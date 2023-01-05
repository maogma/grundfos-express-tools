import datetime
import pandas as pd
import os
import numpy as np
from _types._types import FilePath
from pandas._typing import FilePath, ReadBuffer
from typing import Union, ParamSpec, Callable
from utils.file_ops import read_files_in_dir
from utils.Dataframe_tools import PSD_BOM_Updates
from pandas import DataFrame
from openpyxl import load_workbook, worksheet, Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, PatternFill

###############################################################################
"""  Finds number of rows in a PSD by finding [END] in column A """


def findLastRow(sheet):

    for myRow in range(1, sheet.max_row + 1):
        for column in "A":
            cell_name = "{}{}".format(column, myRow)
            if sheet[cell_name].value == "[END]":
                return myRow

###############################################################################

###############################################################################


def find_specific_cell(searchValue, row_or_col, cols="ABCDEFGHIJKLMNOPQRSTUVWXYZ"):
    # print("looking for {}, type: {}".format(searchValue,type(searchValue)))

    last_row = findLastRow(dimSheet)

    for eachRow in range(1, last_row):
        for column in cols:  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, eachRow)

            # Checking when excel cell value matches searchValue
            if dimSheet[cell_name].value == searchValue:
                # print("cell position {} has value {}".format(cell_name, dimSheet[cell_name].value))
                if row_or_col == "row":
                    return eachRow
                elif row_or_col == "col":
                    return column

##############################################################################


def add_removal_note():
    timeStamp = datetime.datetime.now().strftime("%m-%d-%Y")
    reason = input("Reason for removal: ")
    authority = input("Who authorized/requested this change? ")
    return timeStamp + " " + reason + " per " + authority


def process_sheet(file: tuple[Union[FilePath, ReadBuffer[bytes], bytes], str],
                  sheet_name: str | list[str],
                  removal_note: str | None | bool = None,
                  use_header: bool = False
                  ) -> tuple[str,
                             Union[tuple[DataFrame, int],
                                   tuple[DataFrame, DataFrame, int]],
                             str,
                             str]:
    """
    Params:
    data: Data comes from the binary of the file or the file path. 
    sheet_name: The name of the sheet to process.
    removal_note: Does the sheet have its own removal note or does it share one with the entire workbook (file)."""
    # Check if sheetname is a valid sheet in the file
    data, fname = file
    try:
        if not sheet_name in pd.ExcelFile(data).sheet_names:
            raise ValueError(f"Worksheet named '{sheet_name}' not found")
    except ValueError as err:
        print(err)
        return
    # Need to get the header dataframe and the psd_dataframe
    header, psd_data, length = PSD_BOM_Updates()._get_df(
        io=data, sheet_name=sheet_name, use_header=use_header, header=None)
    # length = tuple of (psd_start_row,original size)
    if not bool(removal_note):
        removal_note = add_removal_note()
    if header:
        return fname, header, psd_data, length, removal_note, sheet_name
    return fname, psd_data, length, removal_note, sheet_name


def process_file(file: FilePath, sheet_list: list[str], output_dir: str, one_removal_note: bool = False):
    # Does this file use the same removal note for every sheet
    if one_removal_note:
        removal_note = add_removal_note()
    else:
        removal_note = None
    if not isinstance(sheet_list, list):
        raise Exception(
            f'You must give the sheets as a list to {process_file.__name__}')
    output_file_name = os.path.join(
        output_dir, "Revised "+os.path.basename(file[1]))
    ret = []
    for sheet in sheet_list:
        val = list(process_sheet(file, sheet, removal_note))
        val.extend([output_file_name])
        ret.append(val)
    return ret


def process_dir(dir, list_sheet_lists, output_dir: str, one_removal_note: bool):
    files = read_files_in_dir(dir)
    ret = []
    for i, file in enumerate(files):
        if file[1].endswith(('.xls', '.xlsx', '.xlsm', '.xlsb', '.odf', '.odt')):
            ret.extend(process_file(
                file, list_sheet_lists[i], output_dir, one_removal_note))
    return ret


def write_new_PSD(file_name,
                  psd_data,
                  grouping_func: Callable[[ParamSpec.args], DataFrame],
                  length: tuple[int, int],
                  removal_note,
                  sheet_name,
                  outPut_file_path
                  ):
    print("Opening file for updates: {}".format(os.path.basename(file_name)))
    psd_startrow = length[0]
    psd_startcol = 0
    end_row = psd_data[psd_data['Full Data'] == '[END]'].index.to_list()[0]
    psd_data = psd_data.iloc[:end_row]
    removals, keep = grouping_func[0](
        psd_data, *grouping_func[1], **grouping_func[2])
    removals.loc[removals["Full Data"] == "[START]", "Full Data"] = ""
    new_row = pd.DataFrame({'ID': removal_note}, index=[0])
    removals = pd.concat([new_row, removals[:]]).reset_index()
    column_list = keep.columns
    removals = removals[column_list]
    removals.sort_values(by=['ID'], inplace=True)
    keep.loc[0, 'Full Data'] = np.nan
    keep.sort_values(by=['ID'], inplace=True)
    keep.reset_index(drop=True, inplace=True)
    keep.loc[0, 'Full Data'] = "[START]"
    num_rows = len(keep)
    keep.loc[num_rows, 'Full Data'] = "[END]"
    append_location = num_rows + psd_startrow
    after_end_row = num_rows + psd_startrow + 2
    wb = load_workbook(file_name)
    ws = wb[sheet_name]
    tabName = sheet_name + "Modified"
    wb.copy_worksheet(ws).title = tabName
    ws_modified = wb[tabName]
    ws_modified.insert_rows(after_end_row-2+len(removals), 2)
    for item in ws_modified.iter_rows(min_col=1, max_col=len(removals.columns), min_row=after_end_row+len(removals)-2, max_row=after_end_row+len(removals)-2):
        for cell in item:
            cell.fill = PatternFill(fill_type=None)
    end_cell = ws_modified.cell(after_end_row-1, column=1)
    end_cell.fill = PatternFill(
        fill_type='solid', start_color="FFCC99", end_color="FFCC99")
    for rows in ws_modified.iter_rows(min_row=append_location+2, max_row=append_location+2, min_col=1, max_col=40):
        for cell in rows:
            cell.fill = PatternFill(
                start_color="FF0000", end_color="FF0000", fill_type="solid")
    if len(removals)+num_rows+psd_startrow == length[1]:
        ws_modified.delete_rows(length[1], len(removals)-1)
    wb.save(outPut_file_path)
    del(ws, ws_modified, wb, end_cell)
    with pd.ExcelWriter(outPut_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        keep.to_excel(writer, sheet_name=tabName, index=False,
                      startrow=psd_startrow-1, startcol=psd_startcol)
        removals.to_excel(writer, sheet_name=tabName, index=False,
                          header=False, startrow=append_location+1, startcol=psd_startcol)
    print("Closing file: {}".format(os.path.basename(file_name)), "\n",
          f"Wrote file: {os.path.basename(outPut_file_path)} to {os.path.dirname(outPut_file_path)}")
