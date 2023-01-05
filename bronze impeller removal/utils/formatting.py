from openpyxl.styles.borders import Border,Side
from openpyxl.styles.fonts import Font
from openpyxl.styles import Alignment,PatternFill
from openpyxl import load_workbook,worksheet,Workbook
from openpyxl.worksheet import cell_range

class BOM_PSD:
    def add_formatting(
        self,
        sheet_name:str,
        outputPath:str,
        new_sheet_name:str
        ):

        psd_startrow=get_start_row(wb,'Impeller',min_row=1,max_col=1)
        
        
        wb=load_workbook(outputPath)
        ws=wb[tabName]
        
        #Adding the Border
        col=ws.column_dimensions['A']
        col.border=Border(right=Side(style='thick'))
        
        #Fitting the Cell Widths in the Columns
        dims = {}
        for row in ws.iter_rows(min_row=0):
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0)+2, len(str(cell.value))+2))
                    if row[0].row <= psd_startrow and cell.column_letter=='A':
                        cell.border=Border(right=Side(style='thick'))
                    elif cell.row<psd_startrow and cell.value=='[END]': #Gets the column end
                        cell.fill=PatternFill(fill_type='solid',start_color="FFCC99",end_color="FFCC99")
        
                        #Applying Top Rows Formatting
                        for row2 in ws.iter_rows(min_col=0,max_col=cell.column-1,min_row=0,max_row=psd_startrow-1):
                            for cell2 in row2:
                                if cell2.row==1:
                                    cell2.fill=PatternFill(fill_type='solid',start_color="808080",end_color="808080")
                                    cell2.font=Font(color='00FFFFFF')
                                    if cell2.column==1:
                                        cell2.font=Font(bold=True,color='00FFFFFF')
                                        cell2.border=Border(right=Side(style='thick'),bottom=Side(style='thick'))
                                    else:
                                        cell2.border=Border(bottom=Side(style='thick'))
                                elif cell2.row==cell.row-1:
                                    cell2.fill=PatternFill(fill_type='solid',start_color="FFCC99",end_color="FFCC99")
                                    if cell2.column==1:
                                        cell2.border=Border(right=Side(style='thick'),bottom=Side(style='thin'))
                                    else:
                                        cell2.border=Border(bottom=Side(style='thin'))
                                elif cell2.row==psd_startrow-1:
                                    cell2.fill=PatternFill(fill_type='solid',start_color="FFCC99",end_color="FFCC99")
                                    if cell2.column==1:
                                        cell2.border=Border(right=Side(style='thick'),bottom=Side(style='thick'))
                                    else:
                                        cell2.border=Border(bottom=Side(style='thick'))
                                else:
                                    cell2.fill=PatternFill(fill_type='solid',start_color="FFCC99",end_color="FFCC99")
                                if cell2.column>1:
                                    cell2.alignment=Alignment(horizontal='center')
                                else:
                                    cell2.alignment=Alignment(horizontal='left')
                        flt=str(cell_range.CellRange(min_col=2, min_row=psd_startrow, max_col=cell.column-1, max_row=psd_startrow))
                        ws.auto_filter.ref=flt     
                    elif cell.row>psd_startrow and cell.column==1 and cell.value=='[START]':
                        cell.fill=PatternFill(fill_type='solid',start_color="FFCC99",end_color="FFCC99")
                        cell.alignment=Alignment(horizontal='right')
                    
                    elif cell.value=='[END]' and cell.row>psd_startrow and cell.column==1:
                        cell.fill=PatternFill(fill_type='solid',start_color="FFCC99",end_color="FFCC99")
                    
                    elif cell.row==append_location+2 and cell.column>=1 and cell.column<=40:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                                           
        for col, value in dims.items():
            ws.column_dimensions[col].width = value
        
        
        wb.save(outputPath)