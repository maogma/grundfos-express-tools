# -*- coding: utf-8 -*-
"""
Created on Tue May 10 12:25:46 2022

@author: 104092
"""
# ###############################################################################
# 1. Load excel
# 2. Find only tabs with "P"
# 3. Count how many diameters (curve tables) there are in Column A
#    a. Make that many duplicate tabs
#    b. Rename tabs to include Diameter
#    c. Delete the unnecessary tables that don't match the diameter for the tab'  
# 4. save to excel file

###############################################################################
def extract_diameters(sh):
    
    list_of_diams = []    
    for row in sh.iter_rows(min_row=10, max_col=1, max_row=21):
        for cell in row:
            if cell.value != None:
                list_of_diams.append(cell.value)
        
    return(list_of_diams)
###############################################################################
def reset_column_A(diameter, sh):
    
    for count, value in enumerate(diameters_list):
        cell_to_change = "{}{}".format('A',10+count)
        cell_to_change2 = "{}{}".format('B',10+count)
        
        if count == 0:
            sh[cell_to_change].value = diameter
            sh[cell_to_change2].value = str(diameter) + '-' + str(diameter)
        else:
            sh[cell_to_change].value = ''
            sh[cell_to_change2].value = ''   
            
    return
###############################################################################
def cleanNewTab(diameter, sh):
    
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter

    max_cols = 21 * len(diameters_list)
    row = 7    
    first_col = 4
    diameter_cols = sorted(list(range(first_col,max_cols,21)), reverse=True)
    
    diameter_cells = []
    
    for item in diameter_cols:
        col = get_column_letter(item)
        cell_coordinate = "{}{}".format(col,row)
        diameter_cells.append(cell_coordinate)        
        
    for cell in diameter_cells:
        xy = coordinate_from_string(cell) # returns ('A',4)
        col_index = column_index_from_string(xy[0]) # returns 1
        # row = xy[1]   
        
        # Check diameter values in row 7.
        # IF diameter doesn't match, delete the following 21 columns
        print("Before Delete - Cell: {} contains {}".format(cell,sh[cell].value))
        if sh[cell].value != diameter:    
            sh.delete_cols(col_index, 21)
            print("After Delete - Cell: {} contains {}".format(cell,sh[cell].value))
        # IF the diameter matches, keep
        else:
            continue

    return
###############################################################################
import os
from openpyxl import load_workbook

myDir = r"C:\Users\104092\OneDrive - Grundfos\Documents\10-19 Projects\12 NBS Curve PSD Separation\12.01 Original Files"
myFile = "GXS Curve_Conexus_V2.xlsx"

filePath = os.path.join(myDir, myFile)

wb = load_workbook(filePath, data_only=True)

for sheet in wb.worksheets:
    if sheet.title[-1] == 'P':
        diameters_list = extract_diameters(sheet)
        print("Diameters in sheet {} are here: {}".format(sheet.title, diameters_list))
        for eachDiameter in diameters_list:    
            tab_to_copy = wb[sheet.title]   
            
           # Add as many tabs as needed per diameter
            print("looking at diameter: {}".format(eachDiameter))
            newTabName = sheet.title + '-D' + str(eachDiameter)
            wb.copy_worksheet(tab_to_copy).title = newTabName
            
            # Remove all additional diameters in column A
            reset_column_A(eachDiameter, wb[newTabName])
            
            # Delete data that doesn't belong in the tab
            cleanNewTab(eachDiameter, wb[newTabName])
        
        wb.remove(sheet)    
        
wb._sheets.sort(key=lambda ws: ws.title)
wb.save(filePath)        



